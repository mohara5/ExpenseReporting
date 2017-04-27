import re
from openpyxl import load_workbook
import win32com.client as win32  
import datetime
from openpyxl.worksheet.datavalidation import DataValidation



def main():
	#The allotment of statments as a text document
	paths = [
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name1.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name2.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name3.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name4.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name5.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name6.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name7.txt', 
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name8.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name9.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name10.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name11.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name12.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name13.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name14.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name15.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name16.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name17.txt', 
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name18.txt',
			'C:\Users\mohara\Documents\ExpenseReports\TESTDOCUMENTS\Name19.txt'
	         ]

	for path in paths:

		dateList = []
		priceList = []
		informationList = []

		#Opening and reading each Statement
		document = open(path, 'r')

		#the Function to extract the name and card number from each statment
		name, card = getCardHolders(path, document)

		#Getting the transactions from each expense report
		transactionList = getTransactions(document)

		for eachTransaction in transactionList:
			eachTransaction = eachTransaction.replace(',','')
			date, price, information = getTransactionData(eachTransaction)
			dateList.append(date)
			priceList.append(price)
			informationList.append(information)

			#Loading in an Expense Reporting excel template to write to
			if len(dateList) == len(priceList) == len(informationList):
				wb = load_workbook('') #insert what workbook template you want to load in
				ws = wb.active
				wb.template = False

				#Automating the Expense Reporting date that this occured
				day = '25' #25th of every month is the day our statements are published. Use your own date
				now = datetime.datetime.now()
				month = now.month
				year = now.year
				year = str(year)
				if month == 1:
					month = 'January'
					lastMonth = 'December'
				elif month == 2:
					month = 'February'
					lastMonth = 'January'
				elif month == 3:
					month = 'March'
					lastMonth = 'February'
				elif month == 4:
					month = 'April'
					lastMonth = 'March'
				elif month == 5:
					month = 'May'
					lastMonth = 'April'
				elif month == 6:
					month = 'June'
					lastMonth = 'May'
				elif month == 7:
					month = 'July'
					lastMonth = 'June'
				elif month == 8:
					month = 'August'
					lastMonth = 'July'
				elif month == 9:
					month = 'September'
					lastMonth = 'August'
				elif month == 10:
					month = 'October'
					lastMonth = 'September'
				elif month == 11:
					month = 'November'
					lastMonth = 'October'
				elif month == 12:
					month = 'December'
					lastMonth = 'November'

				ws['B4'] = str(name)
				ws['B5'] = "Card ending in" + " " + str(card)
				ws['B6'] = "Home Office"
				ws["B7"] = "Statement Period:" + " " + day + " " + lastMonth + " - " + day + " " + month + " " + year
				
				#Example 
				#ws["B7"] = "Statement Period: 25 February - 25 March 2017"

				#Iterating through each column and row in excel to write the information to
				#Writing the date
				i = 0
				column = 'A'
				beginningCell = 9
				for eachDate in dateList:
					i = i+1
					beginningCell = beginningCell + 1
					theCell = str(column) + str(beginningCell)
					ws[theCell] = eachDate
					print eachDate
					print theCell
					print '\n'

				#Writing the Merchant Info
				i = 0
				column = 'B'
				beginningCell = 9
				for eachInformation in informationList:
					i = i+1
					beginningCell = beginningCell + 1
					theCell = str(column) + str(beginningCell)
					ws[theCell] = eachInformation
					print eachInformation
					print theCell
					print '\n'	
					
				#Writing the price	
				i = 0
				column = 'F'
				beginningCell = 9
				for eachPrice in priceList:
					i = i+1
					beginningCell = beginningCell + 1
					theCell = str(column) + str(beginningCell)
					ws[theCell] = eachPrice
					print eachPrice
					print theCell
					print '\n'

				# Naming the specific Expense Report Excel doc and the path to where it is saved
				documentName = 'C:\Folder1\Folder2\\Documents\Folder3\Folder4\/'+'Expense Report'+ ' ' + name + ' ' + card + '.xlsx'
				wb.save(documentName)

	
			else:
				print "THERE IS AN ERROR IN TRANSACTIONS"

		
		#Email specifics
		text = "Please verify that all charges on your expense report are valid and correct. If correct, please sign and return to Accounting and Finance Department"
		subject = "Monthly Expense Report" 
		
		#Who you want to send the email to
		recipient = "" #Place recipients email address inside the ""
		
		#Attaching the Excel sheet for the specific individual we are sending it to
		attachment = documentName

		#Send email
		Emailer(text, subject, recipient, attachment)



#Find Cardholder's Name & Last digits of card
def getCardHolders(path, document):
	cardHolders = {
			   'Name1':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name2':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name3':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name4':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name5':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name6':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name7':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name8':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name9':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name10':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name11':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name12':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name13':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name14':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name15':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name16':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name17':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name18':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name19':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name20':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name21':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name22':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name23':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name24':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name25':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name26':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name27':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name28':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name29':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name30':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name31':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name32':'Card Ending ****', #Place the last 4 numbers of the CC into the stars
			   'Name33':'Card Ending ****' 

			      }

	for name, card in cardHolders.viewitems():
		if name in open(path).read() and card in open(path).read():
			return name, card

#Get each transaction line
def getTransactions(document):
	transactionList = []
	for newLine in document:
		potentialTransaction = re.match(r'([0-9][0-9]-[0-9][0-9] [0-9][0-9]-[0-9][0-9] )', newLine)
		if potentialTransaction:
			transactionList.append(newLine)
	return transactionList


#Get Data from each transaction
def getTransactionData(eachTransaction):

	date = re.findall(r'^([0-9][0-9]-[0-9][0-9])', eachTransaction)
	price = re.findall(r'([0-9]*\.+[0-9]+[0-9]+\s[C][R])|([0-9]*\.+[0-9]+[0-9])', eachTransaction)
	information = re.findall(r'[0-9][0-9]-[0-9][0-9] [0-9][0-9]-[0-9][0-9] (.*?) [0-9]*\.[0-9]*|[0-9]*[\,]*[0-9]*\.[0-9]*', eachTransaction)


	if price[0][0]:
		price = '-'+ price[0][0].replace("CR","")
		return date[0], float(price), information[0]
	else:
		return date[0], float(price[0][1]), information[0]

#Automated email generation
def Emailer(text, subject, recipient, attachment):  

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.HtmlBody = text
    mail.Attachments.Add(Source = attachment)
    mail.Display(True)


if __name__ == '__main__':
	main()